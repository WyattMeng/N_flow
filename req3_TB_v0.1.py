# -*- coding: utf-8 -*-
"""
Created on Tue Oct 10 11:31:41 2017

@author: Maddox.Meng
"""

'''Req3: "A3_WP.xlsx">>>"TB">>>K -------->  "N_WP.xlsx">>>N100>>>TB'''

import os
from openpyxl import load_workbook

import tkFileDialog

import tkMessageBox
tkMessageBox.showinfo( "Hint:", 'Please select\n2016 A3 WP.xlsx first,\n2016 N WP.xlsx  then.')


default_dir = r"%USERPROFILE%\Desktop"  # 设置默认打开目录
A3_WP = tkFileDialog.askopenfilename(title=u"选择文件",
                                   initialdir=(os.path.expanduser(default_dir)))
N_WP = tkFileDialog.askopenfilename(title=u"选择文件",
                                  initialdir=(os.path.expanduser(default_dir)))


from Tkinter import *
import Tkinter as tk
root = tk.Tk()
scrollbar = Scrollbar(root)
scrollbar.pack( side = RIGHT, fill=Y )
text = Text(root, yscrollcommand = scrollbar.set)
text.configure(font=("微软雅黑", 10))
scrollbar.config( command = text.yview )
def scrollwheel(event):
    text.yview_scroll(-1*(event.delta/120), "units")
text.bind_all('<MouseWheel>',scrollwheel)

def tkMessage(textWidget, textContent):
    textWidget.insert(INSERT, textContent)
    textWidget.update()
    textWidget.see('end')  


'''========================================================================='''

tkMessage(text, '| -- '+N_WP.split('/')[-1]+'\n')
text.pack()

wb = load_workbook(A3_WP,  data_only=True) #A3_WP, data_only=True


if 'TB' not in wb.sheetnames:
    print 'Error: No sheet TB in A3_WP'
else:
    
    ws = wb.get_sheet_by_name('TB')
    
#找到“科目号” 
for x in range(1, ws.max_row+1):
    for y in range(0, ws.max_column):
        if ws[x][y].value == u'科目号':
            #print x,y,ws[x][y].value
            subjRow = x
            subjCol = y
            break_flag = True
            break  #这里要加break，否则行数太多，会不停检索到最后一行
    if break_flag == True:
        break        

#找到“Mapping” 
for x in range(1, ws.max_row+1):
    for y in range(0, ws.max_column):
        if ws[x][y].value == 'Mapping':
            #print x,y,ws[x][y].value
            mapping_col = y+1
            break_flag = True
            break  #这里要加break，否则行数太多，会不停检索到最后一行 
    if break_flag == True:
        break 

            
#dict = {}
#subjCodes = []
dict2 = {}
k = 0
for x in range(subjRow+1, ws.max_row+1):
    if ws[x][subjCol].value is not None:
        #dict[k] = {'subjNo': ws[x][subjCol].value, 'value': ws[x][mapping_col].value}
        #subjCodes.append(ws[x][subjCol].value)
        dict2[ws[x][subjCol].value] = ws[x][mapping_col].value
        print ws[x][mapping_col].value
        k+=1  
         
#print len(dict) #620
  

'''============================================================================='''
N_sheets = ['N100', 'N200', 'N300']

wb_NWP = load_workbook(N_WP)
for sheet in N_sheets:
    if sheet in wb_NWP.sheetnames:
        
        
        tkMessage(text, '    | -- '+sheet+'\n')
        print sheet
        
        ws = wb_NWP.get_sheet_by_name(sheet)
        
        
        #找到“Check Mapping” 
        break_flag=False
        for x in range(1, ws.max_row+1):
            for y in range(0, ws.max_column):
                if ws[x][y].value == 'Check Mapping':
                    print x,y,ws[x][y].value         #29 0 Check Mapping
                    
                    chkMap_x_min = x
                    chkMap_y = y
        
                    break_flag = True
                    break
            if break_flag == True:
                break  

        #找到check mapping的area最后一行，
        y = chkMap_y
        for x in range(chkMap_x_min+1, ws.max_row+1):
            if ws[x][y].value != None:
                print ws[x][y].value
                
                chkMap_x_max = x - 1
                
                break
            else: #如果check mapping以下都是None，那么最大行就是ws最大行
                
                chkMap_x_max = ws.max_row
        
        #找到‘TB’、‘科目码’的x，y
        break_flag=False
        for x in range(chkMap_x_min, chkMap_x_max+1):
            for y in range(0, ws.max_column):
                    if ws[x][y].value == 'TB':  #31 3 TB
                        print x,y,ws[x][y].value
                        
                        TB_x = x
                        TB_y = y
                        
                    elif isinstance(ws[x][y].value, (int, long, float)) :   
                        print x,y,ws[x][y].value
                        
                        subjCode_x_min = x
                        subjCode_y     = y
                        
                        break_flag = True
                        break
            if break_flag == True:
                break 



        y = subjCode_y    
        for x in range(TB_x+1, chkMap_x_max+1):
            if ws[x][y].value in dict2:
                print 'code = ',ws[x][y].value,type(ws[x][y].value), 'value = ',dict2[ws[x][y].value]
                tkMessage(text, '        | -- Subject Code : '+ str(ws[x][y].value) +'\n')
                #text.pack()
                #在TB列写入
                ws[x][TB_y].value = dict2[ws[x][y].value]
                
                
wb_NWP.save(N_WP)                
                
tkMessage(text, '\n ========Write Successfully!=========\n') 
root.mainloop()                
'''
python跳出多层循环节约时间
读不到 
'''           
