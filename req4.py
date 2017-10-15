# -*- coding: utf-8 -*-
"""
Created on Sat Oct 14 23:23:55 2017

@author: mmmkk
"""

'''req4'''

import pandas as pd
import numpy as np
import datetime

wp = '重庆市沙坪坝融兴村镇银行有限责任公司-N-20151231-WP.xlsx'.decode('utf8')
wp2 = '重庆市沙坪坝融兴村镇银行有限责任公司-N-20161231-WP.xlsx'.decode('utf8')

def dateToStr(date):
    return  date.strftime('%Y-%m-%d')



#df = pd.read_excel(wp, sheetname='N100', headers=None)
#
##找到2015年的cell的col，组成list
#yearCols = []
#for x in range(0, df.shape[0]):
#    for y in range(0, df.shape[1]):
#        if isinstance(df.iloc[x,y], pd.Timestamp):
#            #print type(df.iloc[x,y].date()) #'datetime.date'
#            if df.iloc[x,y].year == 2015:
#                yearCols.append(y)
#
#
##找到‘审定数据’，如果col在yearCols里，就记录
#for x in range(0, df.shape[0]):
#    for y in range(0, df.shape[1]):
#        if unicode(df.iloc[x,y]) == u'审定数据' and y in yearCols:
#        #UnicodeWarning: Unicode equal comparison failed to convert both arguments 
#        #to Unicode - interpreting them as being unequal #solve: unicode(df.iloc[x,y])    
#            print x, y, df.iloc[x,y]
#            data_y = y
#            x_min = x + 1
#
##找到‘Check Mapping’，作为x_max
#for x in range(0, df.shape[0]):
#    for y in range(0, df.shape[1]):
#        if unicode(df.iloc[x,y]) == u'Check Mapping':
#            print x, y, df.iloc[x,y]
#            x_max = x - 1
#
#            
##找到2、3列中，审定数据下一行，一直到‘合计’那行的区域里，不为空的cell
#for x in range(x_min, x_max+1):
#    for y in [1,2]:
#        if unicode(df.iloc[x,y]) == u'合计':
#            print x, y, df.iloc[x,y]
#            
#            subj_y = y
#            subj_x_max = x
#            
##在subj_y列中搜审定数据下一行，一直到‘合计’那行的区域里，不为空的cell
#dict = {} 
#k=0           
#y = subj_y            
#for x in range(x_min, x_max+1):
#    '''这里判断需要更多实例确保覆盖所有情况'''
#    #if isinstance(df.iloc[x,y], [np.float, np.nan]) is False:
#    if isinstance(df.iloc[x,y], unicode):    
#        print x, y, df.iloc[x,y]
#        
#        dict[k] = {'subj':df.iloc[x,y], 'value': df.iloc[x,data_y]}
#        k+=1            


from openpyxl import load_workbook
wb = load_workbook(wp2,data_only = True)
ws = wb.get_sheet_by_name('N100')

yearCols2 = []
for x in range(1, ws.max_row+1):
    for y in range(0, ws.max_column):
        if isinstance(ws[x][y].value, datetime.datetime) and ws[x][y].value.year==2015:
            #print x, y, ws[x][y].value.year
            yearCols2.append(y)
            break
        
for x in range(1, ws.max_row+1):
    for y in range(0, ws.max_column):
        if ws[x][y].value == u'审定数据' and y in yearCols2:
            print x, y, ws[x][y].value
            
            data_y = y
            x_min = x + 1
            
for x in range(1, ws.max_row+1):
    for y in range(0, ws.max_column):
        if ws[x][y].value == u'Check Mapping':
            print x, y, ws[x][y].value
            
            x_max = x - 1

for x in range(x_min, x_max + 1):
    for y in range(0, ws.max_column):
        if ws[x][y].value == u'合计':
            print x, y, ws[x][y].value
            
            subj_y = y
            subj_x_max = x            

y = subj_y
for x in range(x_min, x_max+1):
    if isinstance(ws[x][y].value, unicode):
        print x, y, ws[x][y].value


#for cell in ws[8]:
#    print cell.value.year, type(cell.value) #2016-12-31 00:00:00 <type 'datetime.datetime'>
    





            