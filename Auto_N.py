# -*- coding: utf-8 -*-
"""
Created on Tue Oct 10 10:14:02 2017

@author: Maddox.Meng
"""

'''Req2: "A3_PBC.xlsx">>>"工作表1">>>ABDE -------->  "A3_WP.xlsx">>>TB>>>ABDE'''

import os
from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string

path = 'C:\Workspace\AuditAutomation_N\N_workflow'

for root, dirs, files in os.walk(path):
    for file in files:
        if (file.decode('gbk').find('~$') == -1 and 
            file.decode('gbk').find('PBC') != -1 and 
            file.decode('gbk').find('A3') != -1): # eliminate temp excel files
            
            PBC = os.path.join(root,file.decode('gbk'))
            #C:\Workspace\AuditAutomation_N\N_workflow\重庆市沙坪坝融兴村镇银行有限责任公司-N-20161231-PBC.xlsx
            
        if (file.decode('gbk').find('~$') == -1 and 
            file.decode('gbk').find('WP') != -1 and 
            file.decode('gbk').find('A3') != -1 and
            file.decode('gbk').find('2016') != -1): # eliminate temp excel files
            
            WP = os.path.join(root,file.decode('gbk'))
            #C:\Workspace\AuditAutomation_N\N_workflow\重庆市沙坪坝融兴村镇银行有限责任公司-N-20161231-WP.xlsx
            
            
#xl = pd.ExcelFile(PBC) #只有一个sheet：工作表1
wb = load_workbook(WP)
df = pd.read_excel(PBC, header = None)
print df.shape    
#找到“科目号” 
for x in range(0, df.shape[0]):
    for y in range(0, df.shape[1]):
        if df.iloc[x,y] == u'科目号':
            x_min = x
            y_min = y
            print x,y   
            
df_rel = pd.read_excel(open(PBC,'rb'), header = x_min) 


for sheetname_WP in wb.sheetnames:
    
    if sheetname_WP == 'TB':
        print sheetname_WP
        ws = wb.get_sheet_by_name(sheetname_WP)  #打开它为ws           
            
headers = [u'科目号', u'科目名称', u'借方余额', u'贷方余额']

for x in range(1, ws.max_row+1):
    for y in range(0, ws.max_column):
        if ws[x][y].value == u'科目号':
            print x,y,ws[x][y].value
            rownumber = x
        break  #这里要加break，否则行数太多，会不停检索到最后一行

#for header in list(df_rel): 
for header in headers:           
    for cell in ws[rownumber]:
        if cell.value == header:
            print 'cell.vaue = header =',cell.value,cell.row,cell.column 
            data_col = df_rel[header]
            data_wp_col = column_index_from_string(cell.column) - 1
            #print data_wp_col
            
            i=0
            y = data_wp_col
            #for x in range(rownumber+1, ws.max_row+1):
            for x in range(rownumber+1, rownumber+1+len(data_col)):    
                
                ws[x][y].value = data_col[i]
                
                i+=1            
            
wb.save(WP)            