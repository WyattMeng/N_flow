# -*- coding: utf-8 -*-
"""
Created on Fri Oct 13 17:34:46 2017

@author: Maddox.Meng
"""

'''req4：审定数据'''

import pandas as pd
import numpy as np
import datetime

wp = r'.\N_workflow\重庆市沙坪坝融兴村镇银行有限责任公司-N-20151231-WP.xlsx'.decode('utf8')
wp2 = r'.\N_workflow\重庆市沙坪坝融兴村镇银行有限责任公司-N-20161231-WP.xlsx'.decode('utf8')

N_sheets = ['N100', 'N200', 'N300', 'N400']

df = pd.read_excel(wp, sheetname='N100', headers = None)

#def dateToStr(date):
#    return  date.strftime('%Y-%m-%d')



#找到2015年的cell的col，组成list
yearCols = []
for x in range(0, df.shape[0]):
    for y in range(0, df.shape[1]):
        if isinstance(df.iloc[x,y], pd.Timestamp):
            #print type(df.iloc[x,y].date()) #'datetime.date'
            if df.iloc[x,y].year == 2015:
                yearCols.append(y)


#找到‘审定数据’，如果col在yearCols里，就记录
for x in range(0, df.shape[0]):
    for y in range(0, df.shape[1]):
        if unicode(df.iloc[x,y]) == u'审定数据' and y in yearCols:
        #UnicodeWarning: Unicode equal comparison failed to convert both arguments 
        #to Unicode - interpreting them as being unequal #solve: unicode(df.iloc[x,y])    
            print x, y, df.iloc[x,y]
            data_y = y
            x_min = x + 1

#找到‘Check Mapping’，作为x_max
for x in range(0, df.shape[0]):
    for y in range(0, df.shape[1]):
        if unicode(df.iloc[x,y]) == u'Check Mapping':
            print x, y, df.iloc[x,y]
            x_max = x - 1

            
#找到2、3列中，审定数据下一行，一直到‘合计’那行的区域里，不为空的cell
for x in range(x_min, x_max+1):
    for y in [1,2]:
        if unicode(df.iloc[x,y]) == u'合计':
            print x, y, df.iloc[x,y]
            
            suj_y = y
            suj_x_max = x
            
#在suj_y列中搜审定数据下一行，一直到‘合计’那行的区域里，不为空的cell
list_src = []
dict = {} 
k=0           
y = suj_y            
for x in range(x_min, x_max+1):
    '''这里判断需要更多实例确保覆盖所有情况'''
    #if isinstance(df.iloc[x,y], [np.float, np.nan]) is False:
    if isinstance(df.iloc[x,y], unicode):    
        #print x, y, df.iloc[x,y]
        
        dict[k] = {'suj':df.iloc[x,y], 'val': df.iloc[x,data_y]}
        print df.iloc[x,data_y]
        k+=1

        list_src.append(df.iloc[x,y])  


list_new = []
i=0        
for item in list_src:
    
    if item is not None and list_src.count(item) == 1:
        list_new.append(item)
    
    if item is not None and list_src.count(item) > 1:
        #print item,list.count(item), i, list[i-1]
        
        for k in range(1, i+1):
            #print i, k, i-k
            if list_src.count(list_src[i-k]) == 1:
                list_new.append(list_src[i-k].strip() + list_src[i].strip())
                break
            else:
                pass#list_new.append(item)
                
    i+=1


for k in dict:
    dict[k]['suj'] = list_new[k]
          

'''==============================读被写入文件================================'''

#{0: {'suj': u'活期存款：', 'val': np.nan},
# 1: {'suj': u'活期存款：公司客户', 'val': 276688716.5499999},
# 2: {'suj': u'活期存款：个人客户', 'val': 48668341.94000019},
# 3: {'suj': u'活期存款：小计', 'val': 325357058.49000007},
# 4: {'suj': u'定期存款：', 'val': np.nan},
# 5: {'suj': u'定期存款：公司客户', 'val': 515857443.77},
# 6: {'suj': u'定期存款：个人客户', 'val': 128324039.64000002},
# 7: {'suj': u'定期存款：小计', 'val': 644181483.41},
# 8: {'suj': u'保证金存款', 'val': 61371289.65000004},
# 9: {'suj': u'应解汇款', 'val': 590000},
# 10: {'suj': u'合计', 'val': 1031499831.5500002}}

from openpyxl import load_workbook
wb = load_workbook(wp2, data_only = True)
ws = wb.get_sheet_by_name('N100')

yearCols2 = []
for x in range(1, ws.max_row+1):
    for y in range(0, ws.max_column):
        if isinstance(ws[x][y].value, datetime.datetime) and ws[x][y].value.year==2015:
            #print x, y, ws[x][y].value.year
            yearCols2.append(y)
            break
        
for x in range(1, ws.max_row+1):
    for y in range(0, ws.max_column):
        if ws[x][y].value == u'审定数据' and y in yearCols2:
            print x, y, ws[x][y].value
            
            data_y = y
            x_min = x + 1
            
for x in range(1, ws.max_row+1):
    for y in range(0, ws.max_column):
        if ws[x][y].value == u'Check Mapping':
            print x, y, ws[x][y].value
            
            x_max = x - 1

for x in range(x_min, x_max + 1):
    for y in range(0, ws.max_column):
        if ws[x][y].value == u'合计':
            print x, y, ws[x][y].value
            
            suj_y = y
            suj_x_max = x            

print '-----------------------'
list_dest = []
dict2 = {}
y = suj_y
for x in range(x_min, x_max+1):
    #if isinstance(ws[x][y].value, unicode):
    #print x, y, ws[x][y].value
    list_dest.append(ws[x][y].value)
    
    if isinstance(ws[x][y].value, unicode):
        dict2[x] = ws[x][y].value

'''有dict，有目标cell位置了，怎么写入'''


'''解决suj非unique的问题'''
list_new2 = []
i=0        
for item in list_dest:
    if item is not None and list_dest.count(item) == 1:
    
        list_new2.append(item)    
    
    if item is not None and list_dest.count(item) > 1:
        #print item,list.count(item), i, list[i-1]
        
        for k in range(1, i+1):
            #print i, k, i-k
            if list.count(list_dest[i-k]) == 1:
                #print list[i-k]+list[i],i-k
                list_new2.append(list_dest[i-k] + list_dest[i].strip())
                break
                
    i+=1  

i=0
for key in dict2:
    dict2[key] = list_new2[i]
    i+=1

for k2 in dict2:
    #print dict2[k2]
    for k in dict:
        if dict[k]['suj'] == dict2[k2]:
            print k2, data_y , dict[k]['val']
            ws[k2][data_y].value = dict[k]['val']
    
wb.save(wp2)            