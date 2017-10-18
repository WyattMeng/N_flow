# -*- coding: utf-8 -*-
"""
Created on Wed Sep 27 10:04:59 2017

@author: Maddox.Meng
"""

'''L-line'''

import os
from openpyxl import load_workbook
import pandas as pd
#from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string
#from openpyxl.utils.dataframe import dataframe_to_columns

#from  datetime  import  * 
import datetime  

import tkFileDialog

default_dir = r"%USERPROFILE%\Desktop"  # 设置默认打开目录
PBC = tkFileDialog.askopenfilename(title=u"选择文件",
                                     initialdir=(os.path.expanduser(default_dir)))

WP = tkFileDialog.askopenfilename(title=u"选择文件",
                                     initialdir=(os.path.expanduser(default_dir)))

#import win32ui
# 
#dlg= win32ui.CreateFileDialog(1)# 1表示打开文件对话框
#dlg.SetOFNInitialDir('%USERPROFILE%\desktop')# 设置打开文件对话框中的初始显示目录
#dlg.DoModal()
# 
#PBC= dlg.GetPathName()# 获取选择的文件名称
#
#'''----------------------------------------------------------'''
#dlg= win32ui.CreateFileDialog(1)# 1表示打开文件对话框
#dlg.SetOFNInitialDir('%USERPROFILE%\desktop')# 设置打开文件对话框中的初始显示目录
#dlg.DoModal()
# 
#WP= dlg.GetPathName()# 获取选择的文件名称






'''========================================================================='''
def stringToDate(string): 
    #example '2013-07-22 09:44:15+00:00' 
    dt = datetime.datetime.strptime(string, "%Y-%m-%d") 
    #print dt 
    return dt

def intToDays(year):
    return datetime.timedelta(days=365*int(year))

def month(date):
    return int(date.split('-')[1])

def year(date):
    return int(date.split('-')[0])

def monthTotal(date):
    return year(date)*12 + month(date)

'''处理当pandas读到日期格式为pandas._libs.tslib.Timestamp的cell，转成string（）'''
def convtPdTimeToStr(time): # datetime.date(2012, 9, 18)
    if isinstance(time, pd._libs.tslib.Timestamp) is True:
        dateStr = time.date().strftime('%Y-%m-%d')
    elif isinstance(time, unicode) is True or isinstance(time, str):
        dateStr = time
    else:
        dateStr = 'date type is %s' % type(time)
    return dateStr  

def convtPdTimeToDate(time):
    if isinstance(time, pd._libs.tslib.Timestamp) is True:
        dateStr = time.date()
    elif isinstance(time, unicode) is True or isinstance(time, str):
        dateStr = stringToDate(time)
    else:
        dateStr = 'date type is %s' % type(time)
    return dateStr 

#if datetime.date(2016,12,31) > stringToDate('2015-12-31'):
#    print 'yes'

path = 'C:\Workspace\AuditAutomation_L\L_workflow'
#CYstart = datetime.date(2016,12,31)#'2016-12-31'
#CYend   = datetime.date(2015,12,31)#'2015-12-31'

cyStart = stringToDate('2015-12-31')#'2016-12-31'
cyStartStr = '2015-12-31'
cyEnd   = stringToDate('2016-12-31')#'2015-12-31'
cyEndStr = '2016-12-31'
print cyStart + datetime.timedelta(days=365*5)

#for root, dirs, files in os.walk(path):
#    for file in files:
#        if file.decode('gbk').find('~$') == -1 and file.decode('gbk').find('PBC') != -1: # eliminate temp excel files
#            PBC = os.path.join(root,file.decode('gbk'))
#        if file.decode('gbk').find('~$') == -1 and file.decode('gbk').find('WP') != -1:
#            WP = os.path.join(root,file.decode('gbk'))
            
#print PBC.encode('utf-8')
#print WP.encode('utf-8') 

#wb = load_workbook(WP)
#for sheetname in wb.sheetnames:
#    print sheetname
#    ws = wb.get_sheet_by_name(sheetname)
    
    
    
xl = pd.ExcelFile(PBC)

wb = load_workbook(WP)


from Tkinter import *
import Tkinter as tk
root = tk.Tk()
scrollbar = Scrollbar(root)
scrollbar.pack( side = RIGHT, fill=Y )
text = Text(root, yscrollcommand = scrollbar.set)
text.configure(font=("微软雅黑", 12))
scrollbar.config( command = text.yview )
def scrollwheel(event):
    text.yview_scroll(-1*(event.delta/120), "units")
text.bind_all('<MouseWheel>',scrollwheel)



logfile = open('logs.txt', 'w+')

for sheet_name in xl.sheet_names:#PBC的每个sheet L110-无形资产清单  L120-无形资产处置清单 
    print sheet_name
    df = pd.read_excel(open(PBC,'rb'), sheetname=sheet_name, header = None)
    
    logfile.write('  | -- '+sheet_name.encode('utf8')+'\n')
    text.insert(INSERT, '  | -- '+sheet_name+'\n')
    text.update()
    text.see('end')    
    
    
    #找到“序号” 
    for x in range(0, df.shape[0]):
        for y in range(0, df.shape[1]):
            if df.iloc[x,y] == u'序号':
                x_min = x
                y_min = y
                
                
    #df_rel = df.iloc[x_min:df.shape[0], y_min:df.shape[1]] 
    #设置“序号”那行为header           
    df_rel = pd.read_excel(open(PBC,'rb'), sheetname=sheet_name, header = x_min)
    
    '''添加计算列'''
    #df_rel['sum'] = df_rel.apply(lambda x: x.sum(), axis=1)
    #df_rel['净值'] = df_rel.apply(lambda row: row[u'原始成本']-row[u'累计摊销额'], axis=1)
    
    df_rel[u'净值']           = df_rel.apply(lambda row: row[u'原始成本']-row[u'累计摊销额'], axis=1)
    df_rel[u'本年新增标志']    = df_rel.apply(lambda row: 1 if stringToDate(row[u'开始摊销时间']) > cyStart else 0, axis=1) # and (stringToDate(row[u'开始摊销时间']) < CYend
    df_rel[u'年初摊销结束标志']= df_rel.apply(lambda row: 1 if (stringToDate(row[u'开始摊销时间'])  +  intToDays(row[u'总摊销年限'])) < cyStart else 0, axis=1)
    df_rel[u'年末摊销结束标志']= df_rel.apply(lambda row: 1 if (stringToDate(row[u'开始摊销时间'])  +  intToDays(row[u'总摊销年限'])) < cyEnd else 0, axis=1)
    df_rel[u'每月摊销金额']    = df_rel.apply(lambda row: row[u'原始成本']/(row[u'总摊销年限']*12), axis=1)
    
    df_rel[u'本年摊销月份']    = df_rel.apply(lambda row:  13 - month(convtPdTimeToStr(row[u'开始摊销时间'])) if row[u'本年新增标志'] == 1 else 
                                                         ( 0 if row[u'年初摊销结束标志'] == 1 else 
                                                         (  ( convtPdTimeToDate(row[u'开始摊销时间']).replace(year=convtPdTimeToDate(row[u'开始摊销时间']).year+row[u'总摊销年限']) ).month if row[u'年末摊销结束标志'] == 1 else 12 )) 
                                             ,axis=1)
    
    df_rel[u'累计摊销月份']    = df_rel.apply(lambda row: min( monthTotal(cyEndStr) - monthTotal(row[u'开始摊销时间']), row[u'总摊销年限']*12), axis=1)
    
    df_rel[u'EY累计摊销额']    = df_rel.apply(lambda row: row[u'每月摊销金额']*row[u'累计摊销月份'], axis=1)
    df_rel[u'DIFF']           = df_rel.apply(lambda row: row[u'EY累计摊销额']-row[u'累计摊销额'], axis=1)
    df_rel[u'本年摊销金额']    = df_rel.apply(lambda row: row[u'每月摊销金额']*row[u'本年摊销月份'], axis=1)
    df_rel[u'新增抽样标志']    = df_rel.apply(lambda row: 'Y' if (row[u'本年新增标志'] == 1 and row[u'原始成本'] > 10000) else '', axis=1)
    
    df_rel.to_excel('res.xlsx')
    
    
    
    
    
    '''写入WP'''
    '''我决定最佳方案是按列写入，所以先按列读取数据'''
    #遍历headers，get每个header对应的数据，即某一列数据
#    for header in list(df_rel):
#        #print header
#        #print df_rel[header]
#        pass

         
     #在WP众多sheets里选出带LL110、L120的sheet
    for sheetname_WP in wb.sheetnames:
        
        if sheetname_WP in sheet_name:
            print sheetname_WP, "in", sheet_name
            ws = wb.get_sheet_by_name(sheetname_WP)  #打开它为ws
        
    for x in range(1, ws.max_row+1):
        for y in range(0, ws.max_column):
            if ws[x][y].value == u'序号':
                print x,y,ws[x][y].value
                rownumber = x

    logfile.write('    | -- WRITING FOLLOWING COLUMN\n')
    text.insert(INSERT, '    | -- WRITING FOLLOWING COLUMN\n')
                
    for header in list(df_rel):
        
        logfile.write('    | -- '+header.encode('utf8')+'\n')
        text.insert(INSERT, '    | -- '+header+'\n')
        text.update()
        text.see('end')  
        text.pack()

            
        for cell in ws[rownumber]:
            if cell.value == header:
                
                print 'cell.vaue = header =',cell.value,cell.row,cell.column
                data_col = df_rel[header]
                data_wp_col = column_index_from_string(cell.column) - 1
                
                i=0
                y = data_wp_col
                #for x in range(rownumber+1, ws.max_row+1):
                for x in range(rownumber+1, rownumber+1+len(data_col)):    
                    
                    ws[x][y].value = data_col[i]
                    
                    i+=1
                    
                    
#                i=rownumber+1    
#                for cell in ws[data_wp_col]:
#                    if i >= rownumber + 1 and (i - rownumber -1) < len(data_col):
#                        print 'i=',i
#                        cell.value = data_col[i-rownumber-1]  #1st x is 8
#                    i+=1
wb.save(WP)           
text.insert(END, 'Successfully!\n')
text.update()
text.see('end') 

root.mainloop()
                
            